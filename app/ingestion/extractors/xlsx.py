from pathlib import Path

from app.ingestion.extractors.base import Extractor


class XlsxExtractor(Extractor):
    supported_extensions = (".xlsx",)

    def extract(self, file_path: Path) -> str:
        from openpyxl import load_workbook

        workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        try:
            sheets: list[str] = []
            for sheet in workbook.worksheets:
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        rows.append(" | ".join(cells))
                if rows:
                    sheets.append(f"# {sheet.title}\n" + "\n".join(rows))
            return "\n\n".join(sheets)
        finally:
            workbook.close()
